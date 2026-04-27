import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import date

st.set_page_config(page_title="CMDB Unos", layout="centered")
st.title("📦 CMDB Unos")

# =========================
# SESSION STATE
# =========================
if "doc_type" not in st.session_state:
    st.session_state.doc_type = None

if "internal_transfer" not in st.session_state:
    st.session_state.internal_transfer = False

# =========================
# DATA
# =========================
DEPLOYMENT_STATES = ["Functional", "Malfunctioned", "Retired"]
INCIDENT_STATES = ["Operational", "Incident"]

TYPE_OPTIONS = [
    "💻 Desktop","💻 Laptop","💵 Cash drawer","📟 Cradle","☎️ IP Phone",
    "🖥️ Monitor","🖥️ Monitor Touch Screen","🧾 Printer Pos","🏷️ Printer label",
    "📡 Router","🔀 Switch","📟 Scanner Counter","✋ Scanner Hand",
    "📱 Scanner Terminal","🔋 UPS","🖧 Server","🖥️ POS Beetle",
    "🖥️ POS Custom","🖥️ POS ELO All in One","🖥️ POS NCR","📦 Other"
]

PROJECTS_MAP = {
    "107 Tendam": "107","108 Deichmann": "108","109 Takko": "109",
    "112 Mercator-S": "112","115 H&M": "115","118 Metre Cash & Carry": "118",
    "119 Ikea": "119","123 Decathlon": "123","193 Lidl": "193"
}

PROJECTS_LABELS = list(PROJECTS_MAP.keys())
UPS_VENDORS = ["APC", "CyberPower", "Socomec", "Inform", "Mustec"]
APC_MODELS = ["APC350", "APC500", "APC650", "APC1000"]

# =========================
# DEVICE INPUT
# =========================
devices = []
valid = True

count = st.number_input("Broj uređaja", 1, 50, 1)

for i in range(int(count)):
    st.markdown("---")
    st.subheader(f"📦 Uređaj {i+1}")

    name = st.text_input("Name *", key=f"name{i}")
    if not name:
        valid = False

    if name == "UPS":
        vendor = st.selectbox("Vendor", [""] + UPS_VENDORS, key=f"vendor{i}")
    else:
        vendor = st.text_input("Vendor", key=f"vendor{i}")

    if vendor == "APC":
        model = st.selectbox("Model", [""] + APC_MODELS, key=f"model{i}")
    else:
        model = st.text_input("Model", key=f"model{i}")

    type_label = st.selectbox("Type *", [""] + TYPE_OPTIONS, key=f"type{i}")
    if not type_label:
        valid = False

    sp = st.text_input("SPInventoryNumber *", key=f"sp{i}")
    sp_clean = sp.strip()

    if not sp_clean or len(sp_clean) != 7 or not (sp_clean.startswith("FS") or sp_clean.startswith("SP")):
        valid = False

    inventory = st.text_input("InventoryNumber", key=f"inv{i}")
    serial = st.text_input("SerialNumber", key=f"serial{i}")

    devices.append({
        "Name": name,
        "Vendor": vendor,
        "Model": model,
        "Type": type_label,
        "SPInventoryNumber": sp_clean,
        "InventoryNumber": inventory,
        "SerialNumber": serial
    })

# =========================
# HELPERS
# =========================
def set_cell(ws, cell, value):
    for merged_range in ws.merged_cells.ranges:
        if cell in merged_range:
            top_left = merged_range.start_cell.coordinate
            ws[top_left] = value
            ws[top_left].alignment = Alignment(horizontal="center", vertical="center")
            return

    ws[cell] = value
    ws[cell].alignment = Alignment(horizontal="center", vertical="center")


def prepare_df():
    df = pd.DataFrame(devices)
    df["Type"] = df["Type"].str.replace(r"[^\w\s\-\/]", "", regex=True).str.strip()
    return df


def validate_devices(df):
    errors = {}

    try:
        existing_df = pd.read_excel("data.xlsx")
    except:
        existing_df = pd.DataFrame()

    for col in ["SPInventoryNumber", "InventoryNumber", "SerialNumber"]:
        if col in existing_df.columns:
            existing_values = set(existing_df[col].astype(str))
            for idx, val in enumerate(df[col]):
                if val and val in existing_values:
                    errors.setdefault(idx, []).append(f"{col} postoji ({val})")

    for col in ["SPInventoryNumber", "InventoryNumber", "SerialNumber"]:
        dup = df[col].duplicated(keep=False)
        for idx in df[dup].index:
            val = df.loc[idx, col]
            if val:
                errors.setdefault(idx, []).append(f"Duplikat ({col}: {val})")

    return errors


def show_errors(errors):
    st.error("❌ Greške:")
    for i, msgs in errors.items():
        st.warning(f"Uređaj {i+1}: " + " | ".join(set(msgs)))


def check(df):
    if not valid:
        st.error("❌ Popuni obavezna polja: Name, Type i SPInventoryNumber")
        st.stop()

    err = validate_devices(df)
    if err:
        show_errors(err)
        st.stop()


@st.cache_data
def load_cmdb_data():
    try:
        return pd.read_excel("data.xlsx", dtype=str).fillna("")
    except:
        return pd.DataFrame()


def find_and_fill_internal(search_col, session_key):
    df = load_cmdb_data()

    if df.empty or search_col not in df.columns:
        return

    value = st.session_state.get(session_key, "").strip()

    if not value:
        return

    match = df[
        df[search_col].astype(str).str.strip().str.upper() == value.upper()
    ]

    if not match.empty:
        row = match.iloc[0]

        st.session_state["int_sp"] = row.get("SPInventoryNumber", "")
        st.session_state["int_inv"] = row.get("InventoryNumber", "")
        st.session_state["int_serial"] = row.get("SerialNumber", "")
        st.session_state["int_name"] = row.get("Name", "")
        st.session_state["int_model"] = row.get("Model", "")

# =========================
# DOCUMENT SELECT
# =========================
st.markdown("---")
col1, col2 = st.columns(2)

if col1.button("📄 Otpremnica"):
    st.session_state.doc_type = "otpremnica"
    st.session_state.internal_transfer = False

if col2.button("📄 Prijemnica"):
    st.session_state.doc_type = "prijemnica"
    st.session_state.internal_transfer = False

# =========================
# OTPREMNICA
# =========================
if st.session_state.doc_type == "otpremnica":
    st.subheader("📄 Otpremnica")

    broj = st.text_input("Broj")
    datum = st.date_input("Datum", value=date.today())
    zaduzio = st.text_input("UREĐAJ ZADUŽIO *")
    objekat = st.text_input("Objekat")
    adresa = st.text_input("Adresa")
    mesto = st.text_input("Mesto")

    if st.button("Generiši Otpremnicu"):
        if not zaduzio.strip():
            st.error("❌ Polje 'UREĐAJ ZADUŽIO' je obavezno")
            st.stop()

        df = prepare_df()
        check(df)

        try:
            wb = load_workbook("otpremnica_template.xlsx")
            ws = wb.active
        except:
            st.error("❌ Nije pronađen fajl: otpremnica_template.xlsx")
            st.stop()

        set_cell(ws, "F4", broj)
        set_cell(ws, "G5", datum.strftime("%d.%m.%Y"))

        set_cell(ws, "G8", zaduzio)
        set_cell(ws, "G9", objekat)
        set_cell(ws, "G10", adresa)
        set_cell(ws, "G11", mesto)

        for i, d in enumerate(devices):
            r = 14 + i
            set_cell(ws, f"B{r}", i + 1)
            set_cell(ws, f"C{r}", d["Name"])
            set_cell(ws, f"D{r}", d["Model"])
            set_cell(ws, f"E{r}", d["InventoryNumber"])
            set_cell(ws, f"F{r}", d["SerialNumber"])
            set_cell(ws, f"G{r}", d["SPInventoryNumber"])

        out = BytesIO()
        wb.save(out)

        st.download_button(
            "Preuzmi Otpremnicu",
            data=out.getvalue(),
            file_name="otpremnica.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# =========================
# PRIJEMNICA
# =========================
if st.session_state.doc_type == "prijemnica":
    st.subheader("📄 Prijemnica")

    broj = st.text_input("Broj")
    datum = st.date_input("Datum", value=date.today())
    magacin = st.text_input("Iz magacina / Ime i prezime *")
    objekat = st.text_input("Objekat")
    adresa = st.text_input("Adresa")
    mesto = st.text_input("Mesto")

    if st.button("Generiši Prijemnicu"):
        if not magacin.strip():
            st.error("❌ Polje 'Iz magacina / Ime i prezime' je obavezno")
            st.stop()

        df = prepare_df()
        check(df)

        try:
            wb = load_workbook("prijemnica_template.xlsx")
            ws = wb.active
        except:
            st.error("❌ Nije pronađen fajl: prijemnica_template.xlsx")
            st.stop()

        set_cell(ws, "F4", broj)
        set_cell(ws, "F5", datum.strftime("%d.%m.%Y"))

        set_cell(ws, "C8", magacin)
        set_cell(ws, "F9", objekat)
        set_cell(ws, "F10", adresa)
        set_cell(ws, "F11", mesto)

        for i, d in enumerate(devices):
            r = 14 + i
            set_cell(ws, f"B{r}", i + 1)
            set_cell(ws, f"C{r}", d["Name"])
            set_cell(ws, f"D{r}", d["SerialNumber"])
            set_cell(ws, f"E{r}", d["SPInventoryNumber"])
            set_cell(ws, f"F{r}", d["InventoryNumber"])

        out = BytesIO()
        wb.save(out)

        st.download_button(
            "Preuzmi Prijemnicu",
            data=out.getvalue(),
            file_name="prijemnica.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# =========================
# INTERNI PRENOS BG -> NS
# =========================
st.markdown("---")
st.subheader("🔁 Interni prenos BG → NS")

if st.button("Interni prenos BG → NS"):
    st.session_state.internal_transfer = True
    st.session_state.doc_type = None

if st.session_state.internal_transfer:

    for key in ["int_sp", "int_inv", "int_serial", "int_name", "int_model"]:
        if key not in st.session_state:
            st.session_state[key] = ""

    st.info("Unesi bilo koji podatak: SP broj, inventarni broj ili serijski broj.")

    st.text_input(
        "SPInventoryNumber",
        key="int_sp",
        on_change=find_and_fill_internal,
        args=("SPInventoryNumber", "int_sp")
    )

    st.text_input(
        "InventoryNumber",
        key="int_inv",
        on_change=find_and_fill_internal,
        args=("InventoryNumber", "int_inv")
    )

    st.text_input(
        "SerialNumber",
        key="int_serial",
        on_change=find_and_fill_internal,
        args=("SerialNumber", "int_serial")
    )

    st.text_input("Name", key="int_name", disabled=True)
    st.text_input("Model", key="int_model", disabled=True)

    if st.button("Preuzmi otpremnicu za interni prenos"):

        if not st.session_state["int_sp"]:
            st.error("❌ Prvo unesi SP / Inventory / Serial da aplikacija pronađe uređaj.")
            st.stop()

        try:
            wb = load_workbook("otpremnica_template.xlsx")
            ws = wb.active
        except:
            st.error("❌ Nije pronađen fajl: otpremnica_template.xlsx")
            st.stop()

        set_cell(ws, "F4", "BG-NS")
        set_cell(ws, "G5", date.today().strftime("%d.%m.%Y"))

        set_cell(ws, "G8", "FSNS")
        set_cell(ws, "G9", "")
        set_cell(ws, "G10", "")
        set_cell(ws, "G11", "")

        r = 14
        set_cell(ws, f"B{r}", 1)
        set_cell(ws, f"C{r}", st.session_state["int_name"])
        set_cell(ws, f"D{r}", st.session_state["int_model"])
        set_cell(ws, f"E{r}", st.session_state["int_inv"])
        set_cell(ws, f"F{r}", st.session_state["int_serial"])
        set_cell(ws, f"G{r}", st.session_state["int_sp"])

        out = BytesIO()
        wb.save(out)

        st.download_button(
            "Preuzmi internu otpremnicu",
            data=out.getvalue(),
            file_name="interni_prenos_BG_NS.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )